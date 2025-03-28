
#Importaciones
from datetime import datetime
import pandas as pd
from  openpyxl import load_workbook
from openpyxl.styles import Font, Alignment, Border, Side
from tqdm import tqdm
import webbrowser
import time
import pyautogui
import json
import zipfile
import os
import shutil



def extraer():
    #Diario
    diario1 = input('Ingrese el diario: ')
    diario = "INDI0" + diario1
    
    #Descargar archivo .zip
    url = 'https://sfg.operations.dynamics.com/?cmp=AGS&mi=DM_DataManagementWorkspaceMenuItem'
    webbrowser.open(url, new=0)
    pyautogui.moveTo(349,507)

    time.sleep(50)
    pyautogui.click()
    pyautogui.moveTo(94,239)
    time.sleep(8)
    pyautogui.click()
    pyautogui.moveTo(311,180)
    time.sleep(50)
    pyautogui.click()
    pyautogui.moveTo(775,420)
    time.sleep(10)
    pyautogui.click()
    time.sleep(10)

    #Buscar archivo
    directorio = 'C:/Users/adan.marchena/Downloads/'
    contenido = os.listdir(directorio)

    nombre_fichero = []
    for fichero in contenido:
        if os.path.isfile(os.path.join(directorio, fichero)) and fichero.endswith('.zip'):
            nombre_fichero.append(fichero)

    print(nombre_fichero)

    #Descomprimir archivo
    print('Descomprimiendo archivo .zip...')
    archivo_zip = f'{directorio}{fichero}'

    ruta = 'C:/Users/adan.marchena/Downloads'
    os.makedirs(ruta, exist_ok=True)

    with zipfile.ZipFile(archivo_zip, 'r') as zip_ref:
        zip_ref.extractall(ruta)


    #Extraer información
    archivo = 'C:/Users/adan.marchena/Downloads/Inventory movement journal headers and lines V4.xlsx'
    hojas = pd.ExcelFile(archivo).sheet_names
    df = pd.DataFrame()
    for hoja in tqdm(hojas, desc='Extrayendo datos de archivo Excel...'):
        df_temp = pd.read_excel(archivo, sheet_name=hoja)
        df = pd.concat([df, df_temp], ignore_index=True)

    print('Organizando datos...')
    #Extraer fecha
    print('Extrayendo fecha actual...')
    fecha = datetime.now()
    fecha_actual = fecha.strftime('%d-%m-%Y')

    #Eliminar filas que no correspondan al almacén 59 y al 105
    print('Eliminando filas no correspondientes...')
    df.drop(df[((df['INVENTORYWAREHOUSEID'] != str(59)) & (df['INVENTORYWAREHOUSEID'] != str(105))) | (df['JOURNALNUMBER'] < diario)].index, inplace=True)

    #Extraer centro de costo
    print('Extrayendo centro de costo...')
    df['CC'] = df['DEFAULTLEDGERDIMENSIONDISPLAYVALUE'].str.extract(r'-(\d+)-')

    #Separar el SALCOS
    print('Especificando tipo de documento...')
    df['TIPODOCUMENTO'] = df['JOURNALNAMEID'].map({'SALCOS':'CONSUMO', 'DEVCOS':'DEVOLUCIÓN'})
    print('Especificando tipo de movimiento...')
    df['TIPOMOVIMIENTO'] = df['JOURNALNAMEID'].map({'SALCOS':'SALIDA', 'DEVCOS':'ENTRADA'})

    #Introducir fechas
    print('Introduciendo fecha actual al DataFrame...')
    df['FECHAACTUAL'] = fecha_actual

    print('Introduciendo fecha del documento al DataFrame...')
    df['TRANSACTIONDATE'] = pd.to_datetime(df['TRANSACTIONDATE'])
    df['TRANSACTIONDATE'] = df['TRANSACTIONDATE'].dt.date
    df['TRANSACTIONDATE'] = pd.to_datetime(df['TRANSACTIONDATE'])

    #Borrar columnas innecesarias
    print('Eliminando columnas innecesarias...')
    df.drop(['PRODUCTCONFIGURATIONID','PRODUCTCOLORID','PRODUCTSIZEID',
            'PRODUCTSTYLEID','INVENTORYSITEID','INVENTORYWAREHOUSEID',
             'ITEMBATCHNUMBER','ITEMSERIALNUMBER','WAREHOUSELOCATIONID',
             'LICENSEPLATENUMBER','INVENTORYSTATUSID','JOURNALNAMEID',
             'INVENTORYOWNERID','LINENUMBER','CATCHWEIGHTQUANTITY','COSTAMOUNT',
             'DEFAULTLEDGERDIMENSIONDISPLAYVALUE','FIXEDCOSTCHARGES',
             'OFFSETMAINACCOUNTIDDISPLAYVALUE','UNITCOST',
             'UNITCOSTQUANTITY'], axis = 'columns', inplace=True)

    #Reemplazar coma por punto para que funcione el float
    print('Cambiando símbolo decimal...')
    df['INVENTORYQUANTITY'] = df['INVENTORYQUANTITY'].str.replace(',', '.').astype(float)
    df['INVENTORYQUANTITY'] = df['INVENTORYQUANTITY'].round(2)

    #Modificar los nombres de las columnas
    print('Modificando nombre de columnas...')
    df.columns = ['diario','cod_producto','cantidad','fecha_documento','CC',
                  'tipo_documento','tipo_movimiento','fecha_actual']

    #Ingresar columnas vacías
    print('Ingresando columnas necesarias...')
    df['num_documento'] = ""
    df['proveedor'] = ""
    df['orden_compra'] = ""
    df['nom_producto'] = ""
    df['detalles'] = ""
    df['cantidad_entrada'] = ""
    df['solicitante'] = ""
    df['receptor'] = ""
    df['vencimiento'] = ""
    df['digitador'] = "ADÁN MARCHENA"

    #Extraer mes y año de la fecha del documento
    print('Separando Mes y año en columnas diferenciadas...')
    df['fecha_documento'] = pd.to_datetime(df['fecha_documento'])
    df['mes'] = df['fecha_documento'].dt.month
    df['año'] = df['fecha_documento'].dt.year
    df['fecha_documento'] = df['fecha_documento'].dt.strftime("%d-%m-%Y")

    #Pasar número a palabra
    print('Cambiando número de mes por nombre de mes...')
    meses = {
        1: 'enero',
        2: 'febrero',
        3: 'marzo',
        4: 'abril',
        5: 'mayo',
        6: 'junio',
        7: 'julio',
        8: 'agosto',
        9: 'septiembre',
        10: 'octubre',
        11: 'noviembre',
        12: 'diciembre'
    }

    df['mes'] = df['mes'].map(meses)

    #Reorganizar las columnas
    print('Reorganizando columnas...')
    df = df.reindex(['fecha_actual','mes', 'año','fecha_documento','tipo_documento',
                     'tipo_movimiento','num_documento','proveedor',
                     'orden_compra','diario','cod_producto','nom_producto',
                     'detalle','cantidad_entrada','cantidad','CC','solicitante',
                     'receptor','vencimiento','digitador'], axis = 1)

    #Pasar la columna cantidad a positivo
    print('Cambiando cantidades negativas a positivas...')
    df['cantidad'] = df['cantidad']*-1

    #Borrar el indi0 del diario
    print('Formatenado el diario...')
    df['diario'] = df['diario'].str.replace('INDI0', '')

    #Llenar columna nombre_producto
    print('Extrayendo información de json...')
    archivo_json = 'C:/Users/adan.marchena/Desktop/Adán Marchena/estudio/scripts/automatizacion/archivos/productos.json'
    with open(archivo_json, 'r', encoding='utf-8') as f:
        datos_json = json.load(f)
    

    diccionario_productos = {item["cod_producto"] : item["nom_producto"] for item in datos_json}

    def obtener_producto(cod_producto):
        return diccionario_productos.get(cod_producto, 'Producto no encontrado')
    
    print('Inyectando información en columna...')
    df['nom_producto'] = df['cod_producto'].apply(obtener_producto)

    #Exportar archivo
    print('Creando archivo Excel...')
    archivo = "Inventario.xlsx"
    df.to_excel(archivo, index=False, engine='openpyxl')

    print('Convirtiendo archivo Excel a DataFrame...')

    libro = load_workbook(archivo)
    hoja = libro.active

    print('Dando formato...')
    fuente = Font(name='Calibri', size=10)
    right_align = Alignment(horizontal='right')
    border = Border(left=Side(border_style='thin'),
                    right=Side(border_style='thin'),
                    top=Side(border_style='thin'),
                    bottom=Side(border_style='thin'))

    for row in hoja.iter_rows():
        for cell in row:
            cell.font = fuente
            cell.border = border
    
    def formatear(columna):
        for cell in hoja[columna]:
            cell.alignment = right_align
            cell.number_format = '0'
    
    formatear('G')
    formatear('J')
    formatear('N')
    formatear('O')
    formatear('P')

    libro.save(archivo)

    #Vaciar carpeta de descargas
    ruta_carpeta = "C:/Users/adan.marchena/Downloads"
    
    for nombre in os.listdir(ruta_carpeta):
        ruta_completa = os.path.join(ruta_carpeta, nombre)
        try:
            if os.path.isfile(ruta_completa):
                os.remove(ruta_completa)
            elif os.path.isdir(ruta_completa):
                shutil.rmtree(ruta_completa)
                
        except Exception as e:
            print(f"Error al eliminar {ruta_completa}: {e}")
    
    print("Carpeta vacía")
        

    print('El archivo ha sido creado con éxito')